import os
import re
import shutil
from openpyxl import load_workbook

from app.core.database import cutoffs_col
from app.models.schemas import PreferenceListItem
from app.services import pdf_service
from app.utils.data import BRANCH_MAPPING


INPUT_XLSX = "scripts/input/students.xlsx"
OUTPUT_DIR = "scripts/output"

os.makedirs(OUTPUT_DIR, exist_ok=True)


def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "", str(name)).strip()


wb = load_workbook(INPUT_XLSX, data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]

col = cutoffs_col()

for row in ws.iter_rows(min_row=2, values_only=True):

    student = dict(zip(headers, row))

    if not student.get("Full Name"):
        continue

    print(f"\nProcessing {student['Full Name']}")

    student_name = str(student["Full Name"]).strip()

    percentile = float(str(student["MHT CET Percentile (Mention Percentile Which Is High)"]).replace("%", "").strip())

    rank = int(str(student["Rank"]).split()[0])

    counselling_id = f"MHCET{rank}"

    mobile = str(student.get("Whatsapp Number", "")).strip()

    # =====================================================
    # EXACTLY SAME AS SEARCH API
    # =====================================================

    seat_types = str(student.get("Category", "")).strip()

    category_list = [
        s.strip()
        for s in seat_types.split(",")
        if s.strip()
    ]

    city_pref = student.get("City Pref")

    city_list = (
        [c.strip() for c in str(city_pref).split(",") if c.strip()]
        if city_pref
        else []
    )

    branch_pref = student.get("Branch Pref")

    selected_general_branches = (
        [b.strip() for b in str(branch_pref).split(",") if b.strip()]
        if branch_pref
        else []
    )

    expanded_branches = []

    for gb in selected_general_branches:
        if gb in BRANCH_MAPPING:
            expanded_branches.extend(BRANCH_MAPPING[gb])
        else:
            expanded_branches.append(gb)

    expanded_branches = list(set(expanded_branches))

    lower_bound = percentile - 30
    upper_bound = percentile + 30

    query = {
        "category": {"$in": category_list},
        "percentile": {
            "$gte": lower_bound,
            "$lte": upper_bound,
        },
    }

    if city_list:
        query["district"] = {"$in": city_list}

    if expanded_branches:
        query["branch_name"] = {"$in": expanded_branches}

    print("Mongo Query:", query)

    docs = (
        col.find(query)
        .sort("percentile", -1)
        .limit(200)
    )

    preference_list = []

    for sr_no, d in enumerate(docs, start=1):
        preference_list.append(
            PreferenceListItem(
                sr_no=sr_no,
                college_code=str(d.get("college_code", "")),
                college_name=d.get("college_name", ""),
                branch_code=str(d.get("branch_code", "")),
                branch_name=d.get("branch_name", ""),
                city=d.get("district", ""),
                category=d.get("category", ""),
                cutoff_percentile=float(d.get("percentile", 0)),
                cutoff_rank=int(d.get("rank", 0)),
            )
        )

    print(f"Found {len(preference_list)} colleges")
    if len(preference_list) == 0:
        break

    pdf_path = pdf_service.fill_template(
        student_name=student_name,
        counselling_id=counselling_id,
        percentile=percentile,
        rank=rank,
        mobile=mobile,
        category=", ".join(category_list),
        preferred_cities=city_list,
        preferred_branches=selected_general_branches,
        preference_list=preference_list,
    )

    final_path = os.path.join(
        OUTPUT_DIR,
        safe_filename(student_name) + ".pdf",
    )

    shutil.move(pdf_path, final_path)

    print(f"Saved -> {final_path}")

print("\nAll PDFs generated successfully.")