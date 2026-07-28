from openpyxl import load_workbook

from app.models.schemas import PreferenceListItem
from app.services import pdf_service

# ==========================
# Student Details
# ==========================
student_name = "Upasana Tarte"
counselling_id = "MHCET4562"
percentile = 94.7805
rank = 22005
mobile = "8080619816"
category = "SBC"

preferred_cities = [
    "Pune",
    "Nashik",
]

preferred_branches = [
    "Computer Engineering",
    "AIML & AIDS",
    "IT",
    "ENTC",
]

# ==========================
# Read Excel
# ==========================
xlsx_file = "scripts/input/Book1.xlsx"

wb = load_workbook(xlsx_file, data_only=True)
ws = wb.active

ordered_list = []

# Read header row
headers = [cell.value for cell in ws[1]]

# Map column names to indexes
col = {name: idx for idx, name in enumerate(headers)}

# Read all data rows
for sr_no, row in enumerate(
    ws.iter_rows(min_row=2, values_only=True),
    start=1
):
    # Skip completely empty rows
    if all(value is None for value in row):
        continue

    ordered_list.append(
        PreferenceListItem(
            sr_no=sr_no,
            college_code=str(row[col["College ID"]]),
            college_name=str(row[col["College Name"]]),
            branch_code=str(row[col["Branch ID"]]),
            branch_name=str(row[col["Branch Name"]]),
            city=str(row[col["City"]]),
            category=str(row[col["Category"]]),
            cutoff_percentile=float(row[col["Percentile"]]),
            cutoff_rank=int(row[col["Rank"]]),
        )
    )

print(f"Loaded {len(ordered_list)} preferences.")

# ==========================
# Generate PDF
# ==========================
pdf_path = pdf_service.fill_template(
    student_name=student_name,
    counselling_id=counselling_id,
    percentile=percentile,
    rank=rank,
    mobile=mobile,
    category=category,
    preferred_cities=preferred_cities,
    preferred_branches=preferred_branches,
    preference_list=ordered_list,
)

print("\nPDF generated successfully!")
print(f"Saved at: {pdf_path}")